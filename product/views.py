from django.forms.models import BaseModelForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.utils import IntegrityError
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from user.permission import IsAdminMixin
from .models import ProductCategory, BudClass, TaxBracket
from .forms import ProductCategoryForm, BudClassForm, TaxBracketForm, ProductPointsForm, ProductPointsUpdateForm
from bill.utils import update_subledger_after_updating_product

class ProductCategoryMixin(IsAdminMixin):
    model = ProductCategory
    form_class = ProductCategoryForm
    paginate_by = 50
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_category_list")
    search_lookup_fields = [
        "title",
        "description",
    ]
class BudClassMixin(IsAdminMixin):
    model = BudClass
    form_class = BudClassForm
    paginate_by = 50
    queryset = BudClass.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_bud_list")
    search_lookup_fields = [
        "title",
        "description",
    ]
class TaxBracketMixin(IsAdminMixin):
    model = TaxBracket
    form_class = TaxBracketForm
    paginate_by = 50
    queryset = TaxBracket.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_taxbracket_list")
    search_lookup_fields = [
        "title",
        "description",
    ]

from .models import ProductPoints
class ProductPointsMixin(IsAdminMixin):
    model = ProductPoints
    form_class = ProductPointsForm
    paginate_by = 50
    queryset = ProductPoints.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_points_list")
    search_lookup_fields = [
        "product__title",
    ]

class ProductPointsList(ProductPointsMixin, ListView):
    template_name = "productpoints/productpoints_list.html"
    queryset = ProductPoints.objects.filter(status=True, is_deleted=False)


class ProductPointsDetail(ProductPointsMixin, DetailView):
    template_name = "productpoints/productpoints_detail.html"


class ProductPointsCreate(ProductPointsMixin, CreateView):
    template_name = "create.html"


class ProductPointsUpdate(ProductPointsMixin, UpdateView):
    form_class = ProductPointsUpdateForm
    template_name = "update.html"


class ProductPointsDelete(ProductPointsMixin, DeleteMixin, View):
    pass


class ProductCategoryList(ProductCategoryMixin, ListView):
    template_name = "productcategory/productcategory_list.html"
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)


class ProductCategoryDetail(ProductCategoryMixin, DetailView):
    template_name = "productcategory/productcategory_detail.html"


class ProductCategoryCreate(ProductCategoryMixin, CreateView):
    template_name = "create.html"


class ProductCategoryUpdate(ProductCategoryMixin, UpdateView):
    template_name = "update.html"


class ProductCategoryDelete(ProductCategoryMixin, DeleteMixin, View):
    pass

class BudClassList(BudClassMixin, ListView):
    template_name = "productbud/productbud_list.html"
    queryset = BudClass.objects.filter(status=True, is_deleted=False)


class BudClassDetail(BudClassMixin, DetailView):
    template_name = "productbud/productbud_detail.html"


class BudClassCreate(BudClassMixin, CreateView):
    template_name = "create.html"


class BudClassUpdate(BudClassMixin, UpdateView):
    template_name = "update.html"


class BudClassDelete(BudClassMixin, DeleteMixin, View):
    pass

class TaxBracketList(TaxBracketMixin, ListView):
    template_name = "producttax/producttax_list.html"
    queryset = TaxBracket.objects.filter(status=True, is_deleted=False)


class TaxBracketDetail(TaxBracketMixin, DetailView):
    template_name = "producttax/producttax_detail.html"


class TaxBracketCreate(TaxBracketMixin, CreateView):
    template_name = "create.html"


class TaxBracketUpdate(TaxBracketMixin, UpdateView):
    template_name = "update.html"


class TaxBracketDelete(TaxBracketMixin, DeleteMixin, View):
    pass



from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import Product
from .forms import ProductForm


class ProductMixin(IsAdminMixin):
    model = Product
    form_class = ProductForm
    paginate_by = 50
    queryset = Product.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductList(ProductMixin, ListView):
    template_name = "product/product_list.html"
    queryset = Product.objects.filter(status=True, is_deleted=False)

# from django.shortcuts import render
# from .models import Product, BudClass

# def product_list_view(request):
#     products = Product.objects.all()
#     budclasses = BudClass.objects.all()
#     context = {
#         'grouped_products': [(budclass.title, budclass.product_set.all()) for budclass in budclasses],
#     }
#     return render(request, 'your_template.html', context)

# from django.views.generic import ListView
# from .models import Product, BudClass

# class ProductList(ListView):
#     template_name = "product/product_list.html"
#     context_object_name = "grouped_products"  # Rename context variable

#     def get_queryset(self):
#         # Fetch all products with specific filters
#         queryset = Product.objects.filter(status=True, is_deleted=False)
#         return queryset

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
        
#         # Fetch all BudClass objects
#         budclasses = BudClass.objects.all()
        
#         # Create a list of tuples with BudClass title and associated products
#         grouped_products = [(budclass.title, budclass.product_set.filter(status=True, is_deleted=False)) for budclass in budclasses]
        
#         # Pass the grouped products to the template
#         context[self.context_object_name] = grouped_products
        
#         return context

from django.views.generic import ListView
from .models import Product, BudClass
from itertools import chain
from django.db.models import Q


# class ProductList(ListView):
class ProductList(ListView, IsAdminMixin):
    template_name = "product/product_list.html"
    context_object_name = "grouped_products"  # Rename context variable

    def get_queryset(self):
        # Fetch all products with specific filters
        queryset = Product.objects.filter(status=True, is_deleted=False)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Fetch all BudClass objects
        budclasses = BudClass.objects.all()

        # If a search query is provided, filter products based on the query
        search_query = self.request.GET.get("q")
        if search_query:
            queryset = Product.objects.filter(Q(title__icontains=search_query, status=True, is_deleted=False)|Q(budclass__title__icontains=search_query, status=True, is_deleted=False)|Q(category__title__icontains=search_query, status=True, is_deleted=False))
            grouped_products = [(budclass.title, queryset.filter(budclass=budclass)) for budclass in budclasses]
        else:
            # Create a list of tuples with BudClass title and associated products
            grouped_products = [(budclass.title, budclass.product_set.filter(status=True, is_deleted=False)) for budclass in budclasses]
        
        # Pass the grouped products to the template
        context[self.context_object_name] = grouped_products
        
        return context

class ProductDetail(ProductMixin, DetailView):
    template_name = "product/product_detail.html"


class ProductCreate(ProductMixin, CreateView):
    template_name = "product/product_create.html"


class ProductUpdate(ProductMixin, UpdateView):
    template_name = "product/product_update.html"

    def form_valid(self, form):
        updated_name = form.data.get('title')
        product_id = form.initial.get('id')
        initial_name = form.initial.get('title')
        update_subledger_after_updating_product(product_id=product_id, initial_name=initial_name, updated_name=updated_name)
        return super().form_valid(form)


class ProductDelete(ProductMixin, DeleteMixin, View):
    pass

from django.shortcuts import redirect
from django.contrib import messages  # Import Django's messaging framework
from .models import TaxBracket, Product
from django.views import View

class ChangeProductTax(View):
    def get(self, request, pk):
        try:
            # Get the TaxBracket object with the specified pk
            tax_bracket = TaxBracket.objects.get(pk=pk)

            # Check if there are products with a null tax bracket
            products_with_tax_true = Product.objects.filter(is_taxable=True)

            if products_with_tax_true.exists():
                # Update the tax bracket for products with a null tax bracket
                products_with_tax_true.update(taxbracket=tax_bracket)
                tax_bracket_name = tax_bracket.title

                # Show a success message
                messages.success(request, f'Taxable Products updated successfully to tax bracket - {tax_bracket_name}')


            else:
                # Show a message indicating that there were no products to update
                messages.info(request, 'No taxable products with tax bracket to update.')

            # Redirect back to the original page or another appropriate URL
            return redirect('product_list')  # Redirect to the product list or change as needed
        except TaxBracket.DoesNotExist:
            # Handle the case where the TaxBracket with the specified pk doesn't exist
            # You can return an error page or redirect to an error page
            pass  # Handle the error as needed


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import CustomerProduct
from .forms import CustomerProductForm


class CustomerProductMixin(IsAdminMixin):
    model = CustomerProduct
    form_class = CustomerProductForm
    paginate_by = 50
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("customerproduct_list")
    search_lookup_fields = ["product__title", "customer__name", "agent__full_name"]


class CustomerProductList(CustomerProductMixin, ListView):
    template_name = "customerproduct/customerproduct_list.html"
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)


class CustomerProductDetail(CustomerProductMixin, DetailView):
    template_name = "customerproduct/customerproduct_detail.html"


class CustomerProductCreate(CustomerProductMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductUpdate(CustomerProductMixin, UpdateView):
    template_name = "update.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductDelete(CustomerProductMixin, DeleteMixin, View):
    pass

'''  STock VIews '''

from .models import ProductStock
from .forms import ProductStockForm

# class ProductStockMixin:
#     model = ProductStock
#     form_class = ProductStockForm
#     paginate_by = 10
#     queryset = ProductStock.objects.filter(status=True,is_deleted=False)
#     success_url = reverse_lazy('productstock_list')
#     search_lookup_fields = [
#         "product__title",
#         "product__description",
#     ]


# class ProductStockList(ProductStockMixin, ListView):
#     template_name = "productstock/productstock_list.html"
#     queryset = ProductStock.objects.filter(status=True,is_deleted=False)
class ProductStockMixin:
    model = ProductStock
    form_class = ProductStockForm
    paginate_by = 10
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('productstock_list')
    search_lookup_fields = [
        "product__title",
        "product__description",
    ]

class ProductStockList(ProductStockMixin, ListView):
    template_name = "productstock/productstock_list.html"

    def get_queryset(self):
        search_query = self.request.GET.get('q')
        queryset = ProductStock.objects.filter(Q(status=True) & Q(is_deleted=False) & ~Q(stock_quantity=0))
        if search_query:
        # Add additional filtering based on the search query
            queryset = queryset.filter(Q(product__title__icontains=search_query) | Q(product__description__icontains=search_query))
        if self.request.GET.get('show_all') == 'true':
            # Show all products regardless of stock_quantity
            queryset = ProductStock.objects.filter(status=True, is_deleted=False)
            if search_query:
            # Add additional filtering based on the search query
                queryset = queryset.filter(Q(product__title__icontains=search_query) | Q(product__description__icontains=search_query))
        return queryset
    
    def get_paginate_by(self, queryset):
        if self.request.GET.get('show_all') == 'true':
            # Show all products, so no pagination needed
            return None
        return self.paginate_by

class ProductStockDetail(ProductStockMixin, DetailView):
    template_name = "productstock/productstock_detail.html"

class ProductStockCreate(ProductStockMixin, CreateView):
    template_name = "create.html"

class ProductStockUpdate(ProductStockMixin, UpdateView):
    template_name = "update.html"

class ProductStockDelete(ProductStockMixin, DeleteMixin, View):
    pass



from django.conf import settings
from django.db import IntegrityError
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from openpyxl import load_workbook
from .models import Product, ProductCategory, BudClass
from urllib.parse import urlparse, urlunparse
import requests
from django.core.files.base import ContentFile

# class ProductUpload(View):

#     def post(self, request):
#         file = request.FILES['file']
#         wb = load_workbook(file)

#         excel_data = list()
#         for sheet in wb.worksheets:
#             for row in sheet.iter_rows():
#                 row_data = list()
#                 for cell in row:
#                     row_data.append(cell.value)
#                 excel_data.append(row_data)
     

#         product_create_error = []
#         for data in excel_data:
#             if not all(data):
#                 continue
#             if data[0].strip().lower().startswith('category'):
#                 continue
#             category_name = data[0].strip().lower().title()

#             if ProductCategory.objects.filter(title__iexact=category_name).exists():
#                 category = ProductCategory.objects.get(title__iexact=category_name)
#             else:
#                 try:
#                     category = ProductCategory.objects.create(title=category_name)
#                 except IntegrityError:
#                     category = ProductCategory.objects.get(title__iexact=category_name)
            
#             product = Product(category=category, title=data[1].strip(), price=float(data[2]), unit=data[3])
#             product.is_taxable = True if data[4].strip().lower() == "yes" else False

#             try:
#                 product.save()
#             except Exception as e:
#                 print(e)
#                 product_create_error.append(product.title)

#         if product_create_error:
#             messages.error(request, f"Error creating products \n {product_create_error}", extra_tags='danger')
#             return redirect(reverse_lazy('product_list'))

#         messages.success(request, "Products uploaded successfully", extra_tags='success')
#         return redirect(reverse_lazy('product_list'))

class ProductUpload(View):

    def post(self, request):
        file = request.FILES['file']
        wb = load_workbook(file)

        excel_data = list()
        # for sheet in wb.worksheets:
        #     for row in sheet.iter_rows():

        #         row_data = list()
        #         for cell in row:
        #             row_data.append(cell.value)
        #         excel_data.append(row_data)

        for sheet in wb.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                # Skip the first row (index 0) which contains column headers
                if row_index == 1:
                    continue

                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                excel_data.append(row_data)

        product_create_error = []
        for data in excel_data:
            if not all(data):
                continue
            title = data[0].strip()
            category_name = data[1].strip().lower()
            price = data[2]
            budclass_name = str(data[2]).strip() if isinstance(data[2], str) else str(data[2]) + " Shelf"
            description = str(data[4]).strip() if isinstance(data[4], str) else str(data[4])
            image = data[3]
            parsed_url = urlparse(str(image))
            image_without_query = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, '', '', ''))

            response = requests.get(image_without_query)

            if ProductCategory.objects.filter(title__iexact=category_name).exists():
                category = ProductCategory.objects.get(title__iexact=category_name)
            else:
                try:
                    category = ProductCategory.objects.create(title=category_name)
                except IntegrityError:
                    category = ProductCategory.objects.get(title__iexact=category_name)

            if BudClass.objects.filter(title__iexact=budclass_name).exists():
                budclass = BudClass.objects.get(title__iexact=budclass_name)
            else:
                try:
                    budclass = BudClass.objects.create(title=budclass_name)
                except IntegrityError:
                    budclass = BudClass.objects.get(title__iexact=budclass_name)

            if response.status_code == 200:
                image_content = ContentFile(response.content)        
                product = Product(
                    category=category,
                    title=title,
                    budclass=budclass,
                    price=price,
                    description=description

                )
                # product.is_taxable = True if data[3].strip().lower() == "yes" else False  # Assuming the 4th column is for is_taxable
                product.image.save(f"{title}_image.png", image_content, save=True)
                try:
                    product.save()
                except Exception as e:
                    print(e)
                    product_create_error.append(product.title)
            else:
                print(f"Failed to download image from {image_without_query}")

        if product_create_error:
            messages.error(request, f"Error creating products \n {product_create_error}", extra_tags='danger')
            return redirect(reverse_lazy('product_list'))

        messages.success(request, "Products uploaded successfully", extra_tags='success')
        return redirect(reverse_lazy('product_list'))

from .models import BranchStock, BranchStockTracking, ItemReconcilationApiItem
from organization.models import Branch, EndDayRecord
from .forms import BranchStockForm
class BranchStockMixin(IsAdminMixin):
    model = BranchStock
    form_class = BranchStockForm
    paginate_by = 10
    queryset = BranchStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('branchstock_list')
    search_lookup_fields = [
        "product__title",
        "product__description",
    ]

class BranchStockList(BranchStockMixin, ListView):
    template_name = "branchstock/branchstock_list.html"

class BranchStockDetail(BranchStockMixin, DetailView):
    template_name = "branchstock/branchstock_detail.html"

class BranchStockCreate(BranchStockMixin, CreateView):
    template_name = "branchstock/branchstock_create.html"

class BranchStockUpdate(BranchStockMixin, UpdateView):
    template_name = "update.html"

class BranchStockDelete(BranchStockMixin, DeleteMixin, View):
    pass

class BranchStockDeleteAll(BranchStockMixin, View):

    def get(self, request, *args, **kwargs):
        # Update the is_deleted field for all BranchStock entries
        BranchStock.objects.all().update(is_deleted=True)
        # Redirect to a success URL
        return redirect('branchstock_list')

from django.db.models import Sum
from bill.models import Bill
from datetime import datetime, date
class ReconcileView(View): 

    def get(self, request):
        opening_exists = BranchStockTracking.objects.count()
        if opening_exists <= 0:
            return render(request, 'item_reconcilation/reconcilation.html',{'show_opening':True})

        branch = request.GET.get('branch', None)
        filter_date = request.GET.get('date')
        branches = Branch.objects.all()
        if not branch or not filter_date:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Please Select a Branch and Date', 'branches':branches})
        try:
            filter_date = datetime.strptime(filter_date, '%Y-%m-%d').date().strftime('%Y-%m-%d')
        except Exception:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Date format must me YYYY-mm-dd', 'branches':branches})

        filter_branch = get_object_or_404(Branch, branch_code__iexact=branch)

        exists_in_bst = BranchStockTracking.objects.filter(date=filter_date, branch=filter_branch).exists()
        if not exists_in_bst:
            products = Product.objects.filter(reconcile=True).order_by('title').values()
            api_items = ItemReconcilationApiItem.objects.filter(date=filter_date, branch=filter_branch, product__reconcile=True).values()
            received = BranchStock.objects.filter(created_at__contains=filter_date, branch=filter_branch, product__reconcile=True).values('product').annotate(quantity=Sum('quantity'))
            bills = Bill.objects.filter(transaction_date=filter_date, branch=filter_branch, status=True)

            new_products = {}
            for product in products:
                for k, v in product.items():
                    if k =='id':
                        physical_count = 0
                        if opening_exists > 0:
                            if BranchStockTracking.objects.filter(branch=filter_branch, product_id=v).exists():
                                physical_count = BranchStockTracking.objects.filter(branch=filter_branch, product_id=v).last().physical
                        new_products[str(v)] = {'title':product.get('title'), 'opening': physical_count}
                        break
        
            for item in api_items:
                product_id = str(item.get('product_id'))
                new_products[product_id]['wastage'] = item.get('wastage', 0)
                new_products[product_id]['returned'] = item.get('returned', 0)
                new_products[product_id]['physical'] = item.get('physical', 0)

            for rec in received:
                product_id = str(rec.get('product'))
                new_products[product_id]['received'] = rec.get('quantity')
            
            for bill in bills:
                for item in bill.bill_items.all():
                    product_id = str(item.product.id)
                    if item.product.reconcile:
                        has_sold = new_products[product_id].get('sold', None)
                        if has_sold:
                            new_products[product_id]['sold'] += item.product_quantity
                        else:
                            new_products[product_id]['sold'] = item.product_quantity

            product_to_view = []
            for k,v in new_products.items():
                new_dict = {'id': k, **v}
                if not 'opening' in new_dict:
                    new_dict['opening'] = 0
                if not 'received' in new_dict:
                    new_dict['received'] = 0
                if not 'wastage' in new_dict:
                    new_dict['wastage'] = 0
                if not 'returned' in new_dict:
                    new_dict['returned'] = 0
                if not 'sold' in new_dict:
                    new_dict['sold'] = 0
                if not 'closing' in new_dict:
                    new_dict['closing'] = 0
                if not 'physical' in new_dict:
                    new_dict['physical'] = 0
                if not 'discrepancy' in new_dict:
                    new_dict['discrepancy'] = 0

                product_to_view.append(new_dict)
            
            for prd in product_to_view:
                opening_received = prd.get('opening') + prd.get('received')
                wastage_returned_sold = prd.get('wastage') + prd.get('returned') + prd.get('sold')
                closing_value = opening_received - wastage_returned_sold
                prd['closing'] = closing_value
                prd['discrepancy'] = prd.get('physical') - closing_value

            context = {
                'products':product_to_view,
                'branches':branches,
                'should_save':True,
                'opening_exists': opening_exists
            }
            return render(request, 'item_reconcilation/reconcilation.html',context)
        
        # --------------------------

        products = BranchStockTracking.objects.filter(date=filter_date, branch=filter_branch).order_by('product__title')
        context = {
            'products':products,
            'branches':branches,
            'should_save':False,
            'opening_exists': opening_exists
        }
        return render(request, 'item_reconcilation/reconcilation.html', context)
    
    def post(self, request):
        branches = Branch.objects.all()
        branch_code = request.POST.get('branch').lower()
        reconcile_date = request.POST.get('filter_date')
        branch = get_object_or_404(Branch, branch_code__iexact=branch_code)
        today_date = date.today()
        if datetime.strptime(reconcile_date, '%Y-%m-%d').date() > today_date:
            messages.error(request, f"Date must not be greater than {today_date}")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})
        
        if BranchStockTracking.objects.filter(date__gte=reconcile_date, branch=branch).exists():
            messages.error(request, f"Items from date greater than {reconcile_date} exists")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})

        last_bill_in_tracking_date = BranchStockTracking.objects.last().date
        bill_exists = Bill.objects.filter(transaction_date__gt=last_bill_in_tracking_date, transaction_date__lt= reconcile_date ,status=True).exists()
        api_items_exists = ItemReconcilationApiItem.objects.filter(date__gt=last_bill_in_tracking_date,date__lt=reconcile_date).exists()

        if bill_exists or api_items_exists:
            messages.error(request, f"Please reconcile items form previous date/s")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})

        data = request.POST
        for k in data:
            try:
                product_id = int(k)
                details = data.getlist(k)
                BranchStockTracking.objects.create(
                    product_id=product_id,
                    branch=branch,
                    date=reconcile_date,
                    opening=details[0],
                    received=details[1],
                    wastage=details[2],
                    returned=details[3],
                    sold=details[4],
                    closing=details[5],
                    physical=details[6],
                    discrepancy=details[7],
                    )
            except ValueError:
                pass
            except IntegrityError:
                messages.error(request, "Items for Today's date already exists")
                return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})
        return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})


class BranchStockUploadView(View):
    
    def post(self, request):
        if BranchStockTracking.objects.count() > 0:
            messages.error(request, "Opening data already exists!!")
            return redirect(reverse_lazy("reconcile"))
        file = request.FILES.get('file')
        branches = Branch.objects.all()
        branch_dict = {}
        for b in branches:
            branch_dict[b.branch_code.lower()] = b.pk

        wb = load_workbook(file)
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    if cell.value:
                        row_data.append(str(cell.value))
                row_data.append(sheet.title)
                excel_data.append(row_data)
       
        product_dict = {}
        for d in excel_data:
            if len(d) < 3:
                continue
            if d[0].lower().startswith('date'):
                continue
            try:
                product_title = d[1].lower().strip()
                product = product_dict.get(product_title, None)
                if not product:
                    product_dict[product_title] = Product.objects.get(title__iexact=product_title).pk
                product_id = product_dict.get(product_title)
                branch_id =  branch_dict.get(d[2].lower())
                # quantity = int(d[2])
                quantity = 0
                opening_date = datetime.strptime(d[0][:10], '%Y-%m-%d')
                BranchStockTracking.objects.create(product_id=product_id, branch_id=branch_id, opening=quantity, physical=quantity, date=opening_date)
            except Exception as e:
                print(e)

        return redirect(reverse_lazy("reconcile"))


class UpdateDateForReconcilationView(View):

    def post(self, request):
        from_date = request.POST.get('from_date', None)
        to_date = request.POST.get('to_date', None)

        if not from_date or not to_date:
            messages.error(request, 'Please Provide both "From date" and "To date"')
            return redirect('/reconcile')
        ItemReconcilationApiItem.objects.filter(date=from_date).update(date=to_date)
        EndDayRecord.objects.filter(date=from_date).update(date=to_date)
        messages.success(request, 'Date has been updated')
        return redirect('/reconcile')
        
class BranchstockUpload(View):

    def post(self, request):
        file = request.FILES['file']
        wb = load_workbook(file)

        excel_data = list()
        for sheet in wb.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                if row_index == 1:
                    continue

                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                excel_data.append(row_data)

        branchstock_create_error = []
        for data in excel_data:
            if not all(data):
                continue
            title = data[0].strip()
            value_to_convert = float(data[2])
            rounded_value = round(value_to_convert)
            stock_count = int(rounded_value)
            branch_id = data[3]
            if Branch.objects.filter(id=branch_id).exists():
                branch = Branch.objects.get(id=branch_id)
            else:
                messages.error(request, f"Error no branch found with {branch_id} \n {branchstock_create_error}", extra_tags='danger')
                return redirect(reverse_lazy('branchstock_list'))

            if Product.objects.filter(title__iexact=title).exists():
                product = Product.objects.get(title=title)

                if stock_count != 0:       
                    branchstock  = BranchStock(
                        branch=branch,
                        product=product,
                        quantity=stock_count
                    )
                try:
                    branchstock.save()
                except Exception as e:
                    print(e)
                    branchstock_create_error.append(product.title)
            else:
                print(f"The product {title} does not exists")

        if branchstock_create_error:
            messages.error(request, f"Error adding products  in branch \n {branchstock_create_error}", extra_tags='danger')
            return redirect(reverse_lazy('branchstock_list'))

        messages.success(request, "Product Stocks uploaded successfully to the branch", extra_tags='success')
        return redirect(reverse_lazy('branchstock_list'))
        
from django.db.models import OuterRef, Subquery, Sum, F
from django.db.models.functions import Coalesce
from django.db.models import Q

# from django.db.models import Q  

# class BranchStockTotalList(View):
#     def get(self, request):
#         search_query = request.GET.get('search')
#         branch_id = request.GET.get('branch')  # Fetch branch_id from GET parameters

#         branches = Branch.objects.all()

#         if search_query:
#             products = Product.objects.filter(
#                 Q(title__icontains=search_query) &
#                 Q(is_deleted=False) &
#                 Q(branchstock__branch_id=branch_id) &
#                 Q(is_menu_item=True)
#             ).annotate(
#                 branchstock_total_quantity=Coalesce(
#                     Subquery(
#                         BranchStock.objects.filter(
#                             product=OuterRef('pk'),
#                             branch_id=branch_id  # Filter based on branch_id
#                         ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]
#                     ),
#                     0
#                 )
#             )
#         else:
#             branchstock_quantity_subquery_all = BranchStock.objects.filter(
#                 product=OuterRef('pk'),
#                 branch=branch_id
#             ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]

#             products = Product.objects.annotate(
#                 branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
#             ).filter(
#                 is_deleted=False,
#                 branchstock_total_quantity__gt=0,
#                 is_menu_item=True
#             )

#         return render(request, 'branchstock/branchstock_total.html', {'branches': branches, 'products': products})

#     def post(self, request):
#         branch_id = request.POST.get('branch')
#         branch = Branch.objects.get(pk=branch_id)

#         branchstock_quantity_subquery_all = BranchStock.objects.filter(
#             product=OuterRef('pk'),
#             branch=branch
#         ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]

#         products = Product.objects.annotate(
#             branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
#         ).filter(
#             is_deleted=False,
#             branchstock_total_quantity__gt=0,
#             is_menu_item=True
#         )

#         return render(request, 'branchstock/branchstock_total.html', {'branches': Branch.objects.all(), 'products': products})

# from django.db.models import Q  

class BranchStockTotalList(View):
    def get(self, request):
        branch_id = request.GET.get('branch')  # Fetch branch_id from GET parameters
        branches = Branch.objects.all()

        branchstock_quantity_subquery_all = BranchStock.objects.filter(
            product=OuterRef('pk'),
            branch_id=branch_id,
            is_deleted=False

        ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]

        products = Product.objects.annotate(
            branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
        ).filter(
            is_deleted=False,
            branchstock_total_quantity__gt=0,
            is_menu_item=True
        )

        return render(request, 'branchstock/branchstock_total.html', {'branches': branches})

    def post(self, request):
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(pk=branch_id)
        search_query = request.POST.get('search')  # Fetch search query from POST parameters

        branchstock_quantity_subquery_all = BranchStock.objects.filter(
            product=OuterRef('pk'),
            branch=branch,
            is_deleted=False

        ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]

        if search_query:
            products = Product.objects.annotate(
                branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
            ).filter(
                title__icontains = search_query,
                is_deleted=False,
                branchstock_total_quantity__gt=0,
                is_menu_item=True
            )
        else:
            products = Product.objects.annotate(
                branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
            ).filter(
                is_deleted=False,
                branchstock_total_quantity__gt=0,
                is_menu_item=True
            )

        return render(request, 'branchstock/branchstock_total.html', {'branches': Branch.objects.all(), 'products': products})


import xlwt
def export_branch_stock(request):
    branch_id = request.GET.get('branch')
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="branch_stock.xls"'

    wb = xlwt.Workbook(encoding="utf-8")
    branches = Branch.objects.all()

    for branch in branches:
        branch_name = branch.name
        ws1 = wb.add_sheet(branch_name)


        branchstock_quantity_subquery_all = (
                    BranchStock.objects.filter(
                        product=OuterRef("pk"), branch_id=branch.id, is_deleted=False
                    )
                    .values("product")
                    .annotate(total_quantity=Sum("quantity"))
                    .values("total_quantity")[:1]
                )

        products = Product.objects.annotate(
                    branchstock_total_quantity=Coalesce(Subquery(branchstock_quantity_subquery_all), 0)
                ).filter(is_deleted=False, is_menu_item=True)


        # Define headers
        headers = ["Title", "Total Quantity"]

        # Write headers to the first row
        for col, header in enumerate(headers):
            ws1.write(0, col, header, xlwt.easyxf("font: bold true"))

        # Write data to subsequent rows
        for row, product in enumerate(products, start=1):
            ws1.write(row, 0, product.title)
            ws1.write(row, 1, product.branchstock_total_quantity)

    ws2 = wb.add_sheet("Inventory")
    products_all = Product.objects.filter(is_deleted=False, status=True)

    headers_1 = ["Product", "Shelf", "Cost Price", "Selling Price", "Category"]

    for col, header in enumerate(headers_1):
            ws2.write(0, col, header, xlwt.easyxf("font:bold true"))

    for row, product in enumerate(products_all, start=1):
            ws2.write(row, 0, product.title)
            ws2.write(row, 1, product.budclass.title)
            ws2.write(row, 2, product.cost_price) 
            ws2.write(row, 3, product.price) 
            ws2.write(row, 4, product.category.title) 

    wb.save(response)
    return response

class ProductThumbnail(View):
    def get(self, request):
        products = Product.objects.filter(is_deleted=False)


        for product in products:
            product.save()

        return redirect('product_list')