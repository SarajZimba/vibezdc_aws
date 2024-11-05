from django.contrib.auth import get_user_model, logout
from django.contrib.auth.models import Group
from django.shortcuts import redirect
from django.urls import reverse_lazy
import requests
import environ
env = environ.Env(DEBUG=(bool, False))

from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin, remove_from_DB
from user.permission import IsAdminMixin

from .forms import UserCreateForm, UserForm, AdminForm


User = get_user_model()


class UserMixin(IsAdminMixin):
    model = User
    form_class = AdminForm
    paginate_by = 50
    queryset = User.objects.filter(status=True)
    success_url = reverse_lazy("user:user_list")
    search_lookup_fields = ["username", "email", "full_name"]


class UserList(UserMixin, ListView):
    template_name = "user/user_list.html"
    queryset = User.objects.filter(is_superuser=True, status=True, is_deleted=False)

    def get_queryset(self, *args, **kwargs):
        queryset = self.queryset.exclude(id=self.request.user.id)
        return queryset


class UserDetail(UserMixin, DetailView):
    template_name = "user/user_detail.html"


class UserCreate(UserMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):
        form.instance.is_superuser = True
        form.instance.is_staff = True
        form.instance.organization = self.request.user.organization

        object = form.save()
        FLASK_URL = env("FLASK_USER_CREATE_URL")
        TOKEN = env("FLASK_USER_CREATE_KEY")
        data= {
            "token":TOKEN,
            "username": object.username,
            "baseURL":self.request.scheme+'://'+self.request.get_host()
        }
        requests.post(
            FLASK_URL,
            json=data
        )
        group = Group.objects.get(name="admin")
        object.groups.add(group)
        return super().form_valid(form)
        


class UserAdmin(UserMixin, CreateView):
    template_name = "create.html"


class UserUpdate(UserMixin, UpdateView):
    template_name = "update.html"

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        old_username = self.object.username
        p = super().post(request, *args, **kwargs)
        FLASK_URL = env("FLASK_USER_UPDATE_URL")
        TOKEN = env("FLASK_USER_CREATE_KEY")
        data= {
            "token":TOKEN,
            "username":old_username,
            "newUsername": request.POST.get('username'),
            "baseURL":request.scheme+'://'+request.get_host(),
            "type":"UPDATE"
        }
        response = requests.post(
            FLASK_URL,
            json=data
        )

        import pdb
        pdb.set_trace()
        return p
        


class UserDelete(UserMixin, View):
    def get(self, request):
        status = remove_from_DB(self, request)
        return JsonResponse({"deleted": status})


def logout_user(request):
    logout(request)
    return redirect(reverse_lazy("user:login_view"))


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import Customer
from .forms import CustomerForm


class CustomerMixin(IsAdminMixin):
    model = Customer
    form_class = CustomerForm
    paginate_by = 50
    queryset = Customer.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("user:customer_list")
    search_lookup_fields = ["name", "tax_number", "contact_number", "email"]


class CustomerList(CustomerMixin, ListView):
    template_name = "customer/customer_list.html"
    queryset = Customer.objects.active()


class CustomerDetail(CustomerMixin, DetailView):
    template_name = "customer/customer_detail.html"


class CustomerCreate(CustomerMixin, CreateView):
    template_name = "customer/customer_create.html"


class CustomerUpdate(CustomerMixin, UpdateView):
    template_name = "update.html"


class CustomerDelete(CustomerMixin, DeleteMixin, View):
    pass


class AgentMixin(IsAdminMixin):
    model = User
    form_class = UserForm
    paginate_by = 50
    queryset = User.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("user:agent_list")
    search_lookup_fields = ["username", "email", "full_name"]


class AgentList(AgentMixin, ListView):
    template_name = "agent/agent_list.html"
    queryset = User.objects.filter(
        groups__name__in=["agent"], status=True, is_deleted=False
    )


class AgentCreate(AgentMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):

        form.instance.is_superuser = False
        form.instance.is_staff = True
        object = form.save()

        group, created = Group.objects.get_or_create(name="agent")
        object.groups.add(group)
        return super().form_valid(form)


class AgentUpdate(AgentMixin, UpdateView):
    template_name = "update.html"


class AgentDelete(AgentMixin, DeleteMixin, View):
    pass

from django.db import IntegrityError
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from openpyxl import load_workbook
from .models import Customer, CustomerNormalLogin
from urllib.parse import urlparse, urlunparse
import requests
from django.core.files.base import ContentFile


# class CustomerUpload(View):

#     def post(self, request):
#         file = request.FILES['file']
#         wb = load_workbook(file)

#         excel_data = list()

#         for sheet in wb.worksheets:
#             for row_index, row in enumerate(sheet.iter_rows(), start=1):
#                 # Skip the first row (index 0) which contains column headers
#                 if row_index == 1:
#                     continue

#                 row_data = list()
#                 for cell in row:
#                     row_data.append(cell.value)
#                 excel_data.append(row_data)

#         product_create_error = []
#         for data in excel_data:
#             if not all(data):
#                 continue
#             # password = data[0].strip()
#             username = data[0]
#             is_active = data[1]
#             name = str(data[2]).strip() if isinstance(data[2], str) else str(data[2]) + str(data[3]).strip() if isinstance(data[3], str) else str(data[3])
#             email = str(data[4]).strip() if isinstance(data[4], str) else str(data[4])
#             mobile_number = data[5]
#             score = data[6]
#             # image_without_query = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, '', '', ''))

#             # response = requests.get(image_without_query)

#             customer = Customer.objects.create(name=name, email = email, loyalty_points=score, contact_number=mobile_number, status=is_active)

#             if customer: 
#                 CustomerNormalLogin.objects.create(customer=customer, username=username)

#                 messages.error(request, f"Error creating products \n {product_create_error}", extra_tags='danger')
#                 return redirect(reverse_lazy('product_list'))

#         messages.success(request, "Products uploaded successfully", extra_tags='success')
#         return redirect(reverse_lazy('product_list'))

from django.contrib import messages

class CustomerUpload(View):

    def post(self, request):
        file = request.FILES['file']
        wb = load_workbook(file)

        excel_data = list()

        for sheet in wb.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                # Skip the first row (index 0) which contains column headers
                if row_index == 1:
                    continue

                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                excel_data.append(row_data)

        customer_create_errors = []
        for data in excel_data:
            # if not all(data):
            #     continue

            username = data[0]
            is_active = data[1]
            # name = str(data[2]).strip() if isinstance(data[2], str) else str(data[2]) + str(data[3]).strip() if isinstance(data[3], str) else str(data[3])
            name = str(data[2]).strip() + " " + str(data[3]).strip() 

            email = str(data[4]).strip() if isinstance(data[4], str) else str(data[4])
            mobile_number = data[5]
            score = data[6]

            try:
                customer = Customer.objects.create(name=name, email=email, loyalty_points=score, contact_number=mobile_number, status=is_active)
                CustomerNormalLogin.objects.create(customer=customer, username=username, email=email)
            except Exception as e:
                customer_create_errors.append(f"Error creating customer: {name}, Error: {str(e)}")

        if customer_create_errors:
            messages.error(request, "\n".join(customer_create_errors), extra_tags='danger')
        else:
            messages.success(request, "Customers uploaded successfully", extra_tags='success')

        return redirect(reverse_lazy('user:customer_list'))
        

class CustomerDeleteAll(CustomerMixin, View):

    def get(self, request, *args, **kwargs):
        # Update the is_deleted field for all BranchStock entries
        Customer.objects.all().update(is_deleted=True)
        # Redirect to a success URL
        return redirect('customer_list')
